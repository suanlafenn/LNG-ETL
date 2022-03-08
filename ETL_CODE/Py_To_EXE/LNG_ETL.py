#!/usr/bin/env python
# coding: utf-8

# In[2]:


### LNG project ETL script
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import os
import datetime
import warnings
warnings.filterwarnings('ignore')

def run_etl(ship_file, price_file, lng_price_file,start,end): 
    """
    Script to do ETL from shippments to data.xlsx for Model to use
    input: shipments file & MTM price file
    output: data.xlsx
    """
    #read data
    
    #DEMM Rate and Boil off
    
    df_price = pd.read_excel(price_file, engine='pyxlsb', sheet_name='Physical', header = 1)
    df_price2 = pd.read_excel(lng_price_file, sheet_name='Daily', header = 2)

    price_rate = df_price2['Shipping Rates'].iloc[-8]

    # price reference 
    df_price= df_price[['Recap No.','Quantity','Average\nPrice ']]
    df_price['Total_Quant'] = df_price.groupby('Recap No.')['Quantity'].transform(sum)
    df_price['price'] = (df_price['Quantity']/df_price['Total_Quant'])*df_price['Average\nPrice ']
    df_price['weighted_price'] = df_price.groupby('Recap No.')['price'].transform(sum)
    df_price = df_price[['Recap No.','weighted_price']]
    df_price.drop_duplicates(subset=['Recap No.', 'weighted_price'],inplace=True)
    df_price.rename(columns={'Recap No.':'REF','weighted_price':'Price'},inplace=True)
    

    wb = load_workbook(ship_file)
    ws = wb["2020-21 Shipments"] # sheet name
    ws.delete_rows(1,3)
    
    #extract the commment of col VPRT
    
    comment_V = []
    for row in ws['V']:
        comment_V.append(row.comment)
    comment_T = []
    for row in ws['T']:
        comment_T.append(row.comment)
    comment_P = []
    for row in ws['P']:
        comment_P.append(row.comment)
    comment_R = []
    for row in ws['R']:
        comment_R.append(row.comment)
    comment_AC = []
    for row in ws['AC']:
        comment_AC.append(row.comment)    
    df = pd.read_excel(ship_file, sheet_name='2020-21 Shipments', header = 2)
    
    
    #take part of data from the file
    df = df[['P/S','LOAD PORT/ Source','INCOTERM','COUNTER PARTY',' VESSEL NAME','Window Starts','Window Ends','LOAD PORT COUNTRY'
            ,'DISPORT COUNTRY','QUANTITY (mmBTU or Cbm)','QUALITY (GHV BTU/SCF &more)','DEMM RATE','BOIL OFF','Trading Chain',
             'CONTRACT NO.','LOAD PORT COUNTRY','DISPORT COUNTRY','DIS PORT']]
    df.rename(columns={'CONTRACT NO.':'REF'},inplace=True)
    
    
    Window=[]
    for index,i in df.iterrows():
        if str(i['DEMM RATE']).startswith('PLATTS'):
            content = price_rate
            Window.append(content)
        else :
            content = i['DEMM RATE']
            Window.append(content)
    df['DEMM RATE'] = Window      

    
    
    
    
    
    #write in the comment 
    df['comment_V'] = comment_V
    df['comment_T'] = comment_T
    df['comment_P'] = comment_P
    df['comment_R'] = comment_R
    df['comment_AC'] = comment_AC
    
    #new file name 
    df.to_excel("output.xlsx") 
    #redecode the file to recoginze the comment data type
    df = pd.read_excel('output.xlsx')

    df['comment_output_p'] = df['comment_P'].str.split('--').str[1]
    df['comment_output_R'] = df['comment_R'].str.split('--').str[1]
    df['comment_output_AC'] = df['comment_AC'].str.split('--').str[1]
    df['comment_output_V'] = df['comment_V'].str.split('--').str[1]
    df['comment_output_T'] = df['comment_T'].str.split('--').str[1]
    

    
    df['Minimum Methane (%)'] = df['comment_output_V'].str.extract(r'((?<=Minimum Methane \(%\):).*)', expand=True)
    df['Maximum Methane (%)'] = df['comment_output_V'].str.extract(r'((?<=Maximum Methane \(%\):).*)', expand=True)
    df['Minimum Ethane (%)'] = df['comment_output_V'].str.extract(r'((?<=Minimum Ethane \(%\):).*)', expand=True)
    df['Maximum Ethane (%)'] = df['comment_output_V'].str.extract(r'((?<=Maximum Ethane \(%\):).*(?=by))', expand=True)
    
    df['Nomination deadline'] = df['comment_output_AC'].astype(str).str.extract(r'((?<=Latest date:).*(?=by))', expand=True)
    
    df['Load_Including'] = df['comment_output_p'].astype(str).str.extract(r'((?<=Including:).*)', expand=True)
    df['Load_Excluding'] = df['comment_output_p'].astype(str).str.extract(r'((?<=Excluding:).*(?=by))', expand=True)
    
    df['DIS_PORT'] = df['comment_output_R'].astype(str).str.extract(r'((?<=Option:).*)', expand=True)
    df['Dis_Including'] = df['comment_output_R'].astype(str).str.extract(r'((?<=Including:).*)', expand=True)
    df['Dis_Excluding'] = df['comment_output_R'].astype(str).str.extract(r'((?<=Excluding:).*(?=by))', expand=True)
    
    df['Tolerance'] = df['comment_output_T'].astype(str).str.extract(r'((?<=Tolerance:).*(?=by))', expand=True)
    
    
    df['index_tolerance'] = df['Tolerance'].str.extract(r'((?<=-).*(?=%))', expand=True)
    df['first_quantity'] = df['QUANTITY (mmBTU or Cbm)'].astype(str).str.extract(r'(.*(?= -))', expand=True)
    df['second_quantity'] = df['QUANTITY (mmBTU or Cbm)'].astype(str).str.extract(r'((?<=- ).*)', expand=True)
    
    
    df['Type'] = 'RE'
    df = df.iloc[:,1:]
    
    #time sort!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    starts =  df.loc[df['Trading Chain'] == start].index.values
    ends =  df.loc[df['Trading Chain'] == end].index.values
    df = df.iloc[int(starts):int(ends)]
    df.dropna(subset=['REF'], inplace=True)
    #merge with price file
    
    df = df.merge(df_price, on=['REF'],how='left')
    
    df['Window Starts'] =pd.to_datetime(df['Window Starts'], errors='coerce',dayfirst=True)
    df['Window Starts'] = pd.to_datetime(df['Window Starts'], format='%Y-%m-%d').dt.strftime('%d/%m/%Y')
    df['Window Ends'] =pd.to_datetime(df['Window Ends'], errors='coerce',dayfirst=True)
    df['Window Ends'] = pd.to_datetime(df['Window Ends'], format='%Y-%m-%d').dt.strftime('%d/%m/%Y')
    Window=[]
    for index,i in df.iterrows():
        if i['Window Starts'] == i['Window Ends']:
            content = i['Window Starts']
            Window.append(content)
        else :
            content = i['Window Starts'] + ' - ' + i['Window Ends']
            Window.append(content)
    df['Delivery/ Loading Window'] = Window
    df['Delivery/ Loading Window'].astype(str)
    
    #cal qty_low. qty_high
    value_low=[]
    value_high=[]
    
    for index,i in df.iterrows():
        if type(i['QUANTITY (mmBTU or Cbm)']) == int:
            cal_val_low = i['QUANTITY (mmBTU or Cbm)'] * (1-float(i['index_tolerance'])*0.01)
            cal_val_high = i['QUANTITY (mmBTU or Cbm)'] * (1+float(i['index_tolerance'])*0.01)
            value_low.append(cal_val_low)
            value_high.append(cal_val_high)
            
        
        else:
            cal_val_low = float(i['first_quantity']) * (1-float(i['index_tolerance'])*0.01)
            cal_val_high = float(i['second_quantity']) * (1+float(i['index_tolerance'])*0.01)
            value_low.append(cal_val_low)
            value_high.append(cal_val_high)
    
    
    df['qty_low'] = value_low
    df['qty_high'] = value_high
    
    #empty columns
    df['Price Description'] = ''
    df['Internal CP'] = ''
    df['SSCS'] = ''
    
    
    df_P = df[df['P/S'] == 'P1']
    df_S = df[df['P/S'] == 'S1']
    
    #sequence and rename for Purchase
    df_P.rename(columns={'REF':'Ref', 'Type':'Type','LOAD PORT/ Source':'Nominated Loading Port',
                        'Load_Including':'Load Including','Load_Excluding':'Load Excluding','DIS_PORT':'Discharge Options',
                        'Dis_Including':'Discharge Including','Dis_Excluding':'Discharge Excluding','INCOTERM':'Incoterm','COUNTER PARTY':'Counter Party',
                        'QUALITY (GHV BTU/SCF &more)':'Spec (GHV)','Minimum Methane (%)':'Min Methane (%)',
                        'Maximum Methane (%)':'Max Methane (%)','Minimum Ethane (%)':'Min Ehtane (%)',
                        'Maximum Ethane (%)':'Max Ethane (%)',' VESSEL NAME':'VESSEL NAME','QUANTITY (mmBTU or Cbm)':'Contractual Quantity',
                        'Tolerance':'Tolerance','Delivery/ Loading Window':'Delivery/ Loading Window','DIS PORT':'Nominated Discharge Port',
                        'qty_low':'qty_low','qty_high':'qty_high','DEMM RATE':'Demm Rate','BOIL OFF':'Boil Off',
                        'LOAD PORT COUNTRY':'Loading Port Country','DISPORT COUNTRY':'Discharge Port Country'},inplace=True)
    #sequence and rename for Sell
    df_S.rename(columns={'REF':'Ref','Delivery/ Loading Window':'Delivery Window','QUANTITY (mmBTU or Cbm)':'Contractual Quantity',
                        ' VESSEL NAME':'VESSEL NAME','QUALITY (GHV BTU/SCF &more)':'Spec (GHV)','DIS PORT':'Nominated Discharge Port',
                        'Minimum Methane (%)':'Min Methane (%)', 'Maximum Methane (%)':'Max Methane (%)','Minimum Ethane (%)':'Min Ehtane (%)',
                        'Maximum Ethane (%)':'Max Ethane (%)','COUNTER PARTY':'Counter Party', 'INCOTERM':'Incoterm',
                        'DIS_PORT':'Discharge Port Option', 'Dis_Including':'Discharge Including','Dis_Excluding':'Discharge Excluding',
                        'LOAD PORT/ Source':'Nominated Loading Port','Load_Including':'Load Including','Load_Excluding':'Load Excluding',
                        'qty_low':'qty_low','qty_high':'qty_high','Type':'Type','DEMM RATE':'Demm Rate','BOIL OFF':'Boil Off',
                        'LOAD PORT COUNTRY':'Loading Port Country','DISPORT COUNTRY':'Discharge Port Country'},inplace=True)
    
    
    
    df_P = df_P[['Ref','Type','Price Description','Price','Nominated Loading Port','Loading Port Country',
                 'Load Including','Load Excluding','Discharge Options','Nominated Discharge Port','Discharge Port Country',
                'Discharge Including','Discharge Excluding','Incoterm','Counter Party','Internal CP','Spec (GHV)','Min Methane (%)',
                'Max Methane (%)','Min Ehtane (%)','Max Ethane (%)','VESSEL NAME','Contractual Quantity',
                'Tolerance','Delivery/ Loading Window','qty_low','qty_high','Demm Rate','Boil Off','Nomination deadline',
                'SSCS']]
    
    df_S = df_S[['Ref','Delivery Window','Contractual Quantity','VESSEL NAME','Spec (GHV)','Min Methane (%)','Max Methane (%)','Min Ehtane (%)',
                'Max Ethane (%)','Counter Party','Internal CP','Incoterm','Discharge Port Country',
                 'Discharge Port Option','Nominated Discharge Port','Discharge Including',
                 'Discharge Excluding','Nominated Loading Port','Loading Port Country',
                'Load Including','Load Excluding','Price','Price Description','qty_low','qty_high','Type','Demm Rate','Boil Off',
                'Nomination deadline','SSCS']]
    
    path = os.path.join(os.path.expanduser("~"), 'Desktop')
    writer = pd.ExcelWriter(path+'\LNG_ETL'+ re.sub(r'[^0-9]','',datetime.datetime.now().strftime("%d%m%Y")) + '.xlsx', engine='xlsxwriter')    
    df_P.to_excel(writer, sheet_name='buy',index=False)
    df_S.to_excel(writer, sheet_name='sell',index=False)
    writer.save()
    os.remove("output.xlsx")
    

    
    
    
def main():
    #os.chdir(os.path.dirname(os.path.abspath(__file__)))
    #print(os.getcwd())

    # ship_file = f"S:/LNG/00 SHIPMENTS/MY SHIPMENTS -20220208.xlsm"
    # price_file = f"S:/LNG/11 MTM/LNG MTM/MTM/2022/MTM - LNG - 08 Feb 2022.xlsb"
    # lng_price_file = f"S:/Common/Platts/LNG/LNG Price 2022.xlsx"

    ship_file = f"MY SHIPMENTS.xlsm"
    price_file = f"MTM-LNG-2022.xlsb"
    lng_price_file = f"LNG Price 2022.xlsx"
    start = 'MAR 22 DATA ABOVE'
    end='APR 22 DATA ABOVE'
    run_etl(ship_file, price_file, lng_price_file,start,end)
    print("Done...")

if __name__ == "__main__":
    main()    


# In[ ]:




